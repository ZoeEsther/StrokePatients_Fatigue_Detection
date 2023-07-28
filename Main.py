import glob
import os
import sys

# import matplotlib
import numpy as np
import openpyxl
# import pandas as pandas
import xlsxwriter
import cv2
import dlib
from PIL import Image
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtChart import QValueAxis, QChart, QSplineSeries, QChartView
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, QComboBox
from PyQt5.QtGui import QIcon, QFont, QImage, QPixmap
from PyCameraList.camera_device import test_list_cameras, list_video_devices, list_audio_devices

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg, NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure

import Utils
import Eye_Detect_Module
import Pose_Estimation_Module
import Fatigue_Scorer_Module

from tqdm import tqdm


# from .Utils import get_face_area
# from .Eye_Detect_Module import EyeDetector as EyeDet
# from .Pose_Estimation_Module import HeadPoseEstimator as HeadPoseEst
# from .Fatigue_Scorer_Module import AttentionScorer as AttScorer

class Main(QMainWindow):
    def __init__(self):
        super().__init__()

        self.threshold = 0.25
        self.use_name = '管理员'
        self.timer_camera = QtCore.QTimer()
        self.timer_lineChart = QtCore.QTimer()
        self.timer_videoSave = QtCore.QTimer()

        self.videoSavePath = None
        self.frameCount = 0
        self.fps = 0
        self.FOURCC = cv2.VideoWriter_fourcc(*'mp4v')

        self.Page1_chart_init()

        self.Data_eye = 'eye'
        self.Data_mouth = 'mouth'
        self.Data_yaw = 'yaw'
        self.Data_roll = 'roll'
        self.Data_pitch = 'pitch'
        self.video_path = ''

        self.DataListEye = []
        self.DataListMouth = []
        self.DataListYaw = []
        self.DataListRoll = []
        self.DataListPitch = []

        self.Detector = dlib.get_frontal_face_detector()
        self.tracker = dlib.correlation_tracker()
        self.Predictor = dlib.shape_predictor(
            r'D:\projects_of_python\StrokePatients_Fatigue_Detection\predictor\shape_predictor_68_face_landmarks.dat')  # instantiation of the dlib keypoint detector model

        self.set_ui()

    def Page1_chart_init(self):
        self.count = 0
        self.lineaxisY1 = QValueAxis()
        self.lineaxisY2 = QValueAxis()
        self.lineaxisY3 = QValueAxis()
        self.lineaxisY4 = QValueAxis()
        self.lineaxisY5 = QValueAxis()
        self.lineaxisX1 = QValueAxis()
        self.lineaxisX2 = QValueAxis()
        self.lineaxisX3 = QValueAxis()
        self.lineaxisX4 = QValueAxis()
        self.lineaxisX5 = QValueAxis()
        self.linechart1 = QChart()
        self.linechart2 = QChart()
        self.linechart3 = QChart()
        self.linechart4 = QChart()
        self.linechart5 = QChart()

        self.lineseries1 = QSplineSeries()
        self.lineseries2 = QSplineSeries()
        self.lineseries3 = QSplineSeries()
        self.lineseries4 = QSplineSeries()
        self.lineseries5 = QSplineSeries()

    def get_Login_Name(self, name):
        self.use_name = name
        self.label_patientName_show.setText(str(self.use_name))
        self.label_patientName_show.repaint()

    def set_ui(self):
        self.setObjectName("MainWindow")
        self.resize(1800, 1150)
        self.centralwidget = QtWidgets.QWidget(self)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.centralwidget.sizePolicy().hasHeightForWidth())
        self.centralwidget.setSizePolicy(sizePolicy)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout_4 = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.TabWidget_MainGUI = QtWidgets.QTabWidget(self.centralwidget)
        self.TabWidget_MainGUI.setObjectName("TabWidget_MainGUI")

        """
        第一个页面，分为两大部分，显示区/曲线区
        """
        self.Camera_page = QtWidgets.QWidget()
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.Camera_page.sizePolicy().hasHeightForWidth())
        self.Camera_page.setSizePolicy(sizePolicy)
        self.Camera_page.setObjectName("Camera_page")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.Camera_page)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setSpacing(5)
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        # 视频播放区
        self.groupBox_Player = QtWidgets.QGroupBox(self.Camera_page)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.groupBox_Player.sizePolicy().hasHeightForWidth())
        self.groupBox_Player.setSizePolicy(sizePolicy)
        self.groupBox_Player.setObjectName("groupBox_Player")
        self.gridLayout = QtWidgets.QGridLayout(self.groupBox_Player)
        self.gridLayout.setObjectName("gridLayout")
        # 尺寸调整带
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout.addItem(spacerItem, 5, 4, 1, 1)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout.addItem(spacerItem1, 5, 0, 1, 1)
        spacerItem2 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem2, 5, 3, 1, 1)
        spacerItem3 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout.addItem(spacerItem3, 5, 6, 1, 1)
        spacerItem4 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem4, 5, 5, 1, 1)
        spacerItem5 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout.addItem(spacerItem5, 5, 2, 1, 1)
        spacerItem6 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem6, 7, 3, 1, 1)
        spacerItem7 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem7, 3, 3, 1, 1)

        self.gridLayout.setRowStretch(0, 3)
        self.gridLayout.setRowStretch(1, 20)
        self.gridLayout.setRowStretch(2, 1)
        self.gridLayout.setRowStretch(3, 1)
        self.gridLayout.setRowStretch(4, 1)
        self.gridLayout.setRowStretch(5, 1)
        self.gridLayout.setRowStretch(6, 1)
        self.gridLayout.setRowStretch(7, 1)
        # 训练状态警示区
        self.label_fatigue_reminder = QtWidgets.QLabel(self.groupBox_Player)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_fatigue_reminder.sizePolicy().hasHeightForWidth())
        self.label_fatigue_reminder.setSizePolicy(sizePolicy)
        self.label_fatigue_reminder.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.label_fatigue_reminder.setObjectName("label_fatigue_reminder")
        self.gridLayout.addWidget(self.label_fatigue_reminder, 0, 0, 1, 7)
        # 播放视频的标签
        self.label_camera_show = QtWidgets.QLabel(self.groupBox_Player)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_camera_show.sizePolicy().hasHeightForWidth())
        self.label_camera_show.setSizePolicy(sizePolicy)
        self.label_camera_show.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.label_camera_show.setObjectName("label_camera_show")
        self.gridLayout.addWidget(self.label_camera_show, 1, 0, 2, 7)
        # 保存按钮
        self.btn_save = QtWidgets.QPushButton(self.groupBox_Player)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.btn_save.sizePolicy().hasHeightForWidth())
        self.btn_save.setSizePolicy(sizePolicy)
        self.btn_save.setObjectName("btn_save")
        self.gridLayout.addWidget(self.btn_save, 6, 5, 1, 1)
        self.btn_save.clicked.connect(self.save)
        # 继续按钮
        self.btn_cont = QtWidgets.QPushButton(self.groupBox_Player)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.btn_cont.sizePolicy().hasHeightForWidth())
        self.btn_cont.setSizePolicy(sizePolicy)
        self.btn_cont.setObjectName("btn_cont")
        self.gridLayout.addWidget(self.btn_cont, 6, 3, 1, 1)
        self.btn_cont.clicked.connect(self.contiPlay)
        # 开始按钮
        self.btn_start = QtWidgets.QPushButton(self.groupBox_Player)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.btn_start.sizePolicy().hasHeightForWidth())
        self.btn_start.setSizePolicy(sizePolicy)
        self.btn_start.setObjectName("btn_start")
        self.gridLayout.addWidget(self.btn_start, 4, 3, 1, 1)
        self.btn_start.clicked.connect(self.start)
        # 停止按钮
        self.btn_stop = QtWidgets.QPushButton(self.groupBox_Player)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.btn_stop.sizePolicy().hasHeightForWidth())
        self.btn_stop.setSizePolicy(sizePolicy)
        self.btn_stop.setObjectName("btn_stop")
        self.gridLayout.addWidget(self.btn_stop, 4, 5, 1, 1)
        self.btn_stop.clicked.connect(self.stop)

        # 选择文件
        self.btn_select = QtWidgets.QPushButton(self.groupBox_Player)
        self.btn_select.setObjectName("btn_select")
        self.gridLayout.addWidget(self.btn_select, 4, 1, 1, 1)
        self.btn_select.clicked.connect(self.selectVideo)

        # 实时拍摄
        self.btn_camera = QtWidgets.QPushButton(self.groupBox_Player)
        self.btn_camera.setObjectName("btn_camera")
        self.gridLayout.addWidget(self.btn_camera, 6, 1, 1, 1)
        self.btn_camera.clicked.connect(self.cameraVideo)

        self.horizontalLayout_2.addWidget(self.groupBox_Player)

        # 曲线图显示区
        self.groupBox_Qchart = QtWidgets.QGroupBox(self.Camera_page)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.groupBox_Qchart.sizePolicy().hasHeightForWidth())
        self.groupBox_Qchart.setSizePolicy(sizePolicy)
        self.groupBox_Qchart.setObjectName("groupBox_Qchart")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.groupBox_Qchart)
        self.gridLayout_3.setObjectName("gridLayout_3")
        # 眼部纵横比
        self.graphicsView_eye = QChartView(self.groupBox_Qchart)
        self.graphicsView_eye.setObjectName("graphicsView_eye")
        self.gridLayout_3.addWidget(self.graphicsView_eye, 0, 0, 1, 1)
        # 嘴部开合比
        self.graphicsView_mouth = QChartView(self.groupBox_Qchart)
        self.graphicsView_mouth.setObjectName("graphicsView_mouth")
        self.gridLayout_3.addWidget(self.graphicsView_mouth, 1, 0, 1, 1)
        # 头部yaw
        self.graphicsView_yaw = QChartView(self.groupBox_Qchart)
        self.graphicsView_yaw.setObjectName("graphicsView_yaw")
        self.gridLayout_3.addWidget(self.graphicsView_yaw, 2, 0, 1, 1)
        # 头部roll
        self.graphicsView_roll = QChartView(self.groupBox_Qchart)
        self.graphicsView_roll.setObjectName("graphicsView_roll")
        self.gridLayout_3.addWidget(self.graphicsView_roll, 3, 0, 1, 1)
        # 头部pitch
        self.graphicsView_pitch = QChartView(self.groupBox_Qchart)
        self.graphicsView_pitch.setObjectName("graphicsView_pitch")
        self.gridLayout_3.addWidget(self.graphicsView_pitch, 4, 0, 1, 1)

        self.horizontalLayout_2.addWidget(self.groupBox_Qchart)
        self.horizontalLayout_2.setStretch(0, 3)
        self.horizontalLayout_2.setStretch(1, 2)
        self.gridLayout_2.addLayout(self.horizontalLayout_2, 0, 0, 1, 1)
        self.TabWidget_MainGUI.addTab(self.Camera_page, "")

        """
        第二个页面，历史数据查询
        """
        self.HistoryData_page = QtWidgets.QWidget()
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.HistoryData_page.sizePolicy().hasHeightForWidth())
        self.HistoryData_page.setSizePolicy(sizePolicy)
        self.HistoryData_page.setObjectName("HistoryData_page")
        self.gridLayout_6 = QtWidgets.QGridLayout(self.HistoryData_page)
        self.gridLayout_6.setObjectName("gridLayout_6")
        self.gridLayout_5 = QtWidgets.QGridLayout()
        self.gridLayout_5.setObjectName("gridLayout_5")

        # 患者名字 标签 不用变
        self.label_patientName = QtWidgets.QLabel(self.HistoryData_page)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_patientName.sizePolicy().hasHeightForWidth())
        self.label_patientName.setSizePolicy(sizePolicy)
        self.label_patientName.setObjectName("label_patientName")
        self.gridLayout_5.addWidget(self.label_patientName, 0, 0, 1, 2)
        # 患者名字显示标签
        self.label_patientName_show = QtWidgets.QLabel(self.HistoryData_page)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_patientName_show.sizePolicy().hasHeightForWidth())
        self.label_patientName_show.setSizePolicy(sizePolicy)
        self.label_patientName_show.setObjectName("label")
        self.gridLayout_5.addWidget(self.label_patientName_show, 0, 2, 1, 1)
        # 历史训练数据 标签 不用变
        self.label_trainNumber = QtWidgets.QLabel(self.HistoryData_page)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_trainNumber.sizePolicy().hasHeightForWidth())
        self.label_trainNumber.setSizePolicy(sizePolicy)
        self.label_trainNumber.setObjectName("label_trainNumber")
        self.gridLayout_5.addWidget(self.label_trainNumber, 1, 0, 1, 2)
        # 历史训练数据选择单选下拉框
        self.comboBox = MyComboBox(self.HistoryData_page)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.comboBox.sizePolicy().hasHeightForWidth())
        self.comboBox.setSizePolicy(sizePolicy)
        self.comboBox.setObjectName("comboBox")
        self.gridLayout_5.addWidget(self.comboBox, 1, 2, 1, 1)
        self.comboBox.insertItem(0, '请选择历史训练数据')
        self.comboBox.clicked.connect(self.renewComboBox)
        # 确定按钮
        self.btn_yes = QtWidgets.QPushButton(self.HistoryData_page)
        self.btn_yes.setObjectName("btn_yes")
        self.gridLayout_5.addWidget(self.btn_yes, 1, 3, 1, 1)
        self.btn_yes.clicked.connect(self.showHisFatigueChart)
        # 活动条
        spacerItem8 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_5.addItem(spacerItem8, 5, 0, 1, 2)
        spacerItem9 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_5.addItem(spacerItem9, 11, 0, 1, 2)
        spacerItem10 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_5.addItem(spacerItem10, 9, 0, 1, 2)
        spacerItem11 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_5.addItem(spacerItem11, 7, 0, 1, 2)
        spacerItem12 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_5.addItem(spacerItem12, 2, 0, 1, 2)
        spacerItem13 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_5.addItem(spacerItem13, 1, 4, 1, 1)
        spacerItem14 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_5.addItem(spacerItem14, 14, 0, 1, 2)
        spacerItem15 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_5.addItem(spacerItem15, 3, 0, 1, 2)
        spacerItem16 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_5.addItem(spacerItem16, 13, 0, 1, 2)

        # 显示曲线的地方
        self.widget = QtWidgets.QWidget(self.HistoryData_page)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.widget.sizePolicy().hasHeightForWidth())
        self.widget.setSizePolicy(sizePolicy)
        self.widget.setObjectName("widget")
        self.gridLayout_5.addWidget(self.widget, 2, 2, 13, 3)

        self.line = MyCanvas()
        self.toolbar = NavigationToolbar(self.line, self.widget)
        self.layout = QtWidgets.QVBoxLayout()

        # 单选按钮 eye
        self.rbtn_eye = QtWidgets.QRadioButton(self.HistoryData_page)
        self.rbtn_eye.setObjectName("rbtn_eye")
        self.gridLayout_5.addWidget(self.rbtn_eye, 4, 0, 1, 2)
        self.rbtn_eye.clicked.connect(self.widget_show_eye)
        # 单选按钮 mouth
        self.rbtn_mouth = QtWidgets.QRadioButton(self.HistoryData_page)
        self.rbtn_mouth.setObjectName("rbtn_mouth")
        self.gridLayout_5.addWidget(self.rbtn_mouth, 6, 0, 1, 2)
        self.rbtn_mouth.clicked.connect(self.widget_show_mouth)
        # 单选按钮 pitch
        self.rbtn_pitch = QtWidgets.QRadioButton(self.HistoryData_page)
        self.rbtn_pitch.setObjectName("rbtn_pitch")
        self.gridLayout_5.addWidget(self.rbtn_pitch, 12, 0, 1, 2)
        self.rbtn_pitch.clicked.connect(self.widget_show_pitch)
        # 单选按钮 roll
        self.rbtn_roll = QtWidgets.QRadioButton(self.HistoryData_page)
        self.rbtn_roll.setObjectName("rbtn_roll")
        self.gridLayout_5.addWidget(self.rbtn_roll, 10, 0, 1, 2)
        self.rbtn_roll.clicked.connect(self.widget_show_roll)
        # 单选按钮 yaw
        self.rbtn_yaw = QtWidgets.QRadioButton(self.HistoryData_page)
        self.rbtn_yaw.setObjectName("rbtn_yaw")
        self.gridLayout_5.addWidget(self.rbtn_yaw, 8, 0, 1, 2)
        self.rbtn_yaw.clicked.connect(self.widget_show_yaw)

        self.gridLayout_5.setColumnStretch(0, 2)
        self.gridLayout_5.setColumnStretch(2, 2)
        self.gridLayout_5.setColumnStretch(3, 1)
        self.gridLayout_5.setColumnStretch(4, 10)
        self.gridLayout_5.setRowStretch(0, 1)
        self.gridLayout_5.setRowStretch(1, 1)
        self.gridLayout_5.setRowStretch(2, 3)
        self.gridLayout_5.setRowStretch(3, 1)
        self.gridLayout_5.setRowStretch(4, 1)
        self.gridLayout_5.setRowStretch(5, 1)
        self.gridLayout_5.setRowStretch(6, 1)
        self.gridLayout_5.setRowStretch(7, 1)
        self.gridLayout_5.setRowStretch(8, 1)
        self.gridLayout_5.setRowStretch(9, 1)
        self.gridLayout_5.setRowStretch(10, 1)
        self.gridLayout_5.setRowStretch(11, 1)
        self.gridLayout_5.setRowStretch(12, 1)
        self.gridLayout_5.setRowStretch(13, 1)
        self.gridLayout_5.setRowStretch(14, 3)
        self.gridLayout_6.addLayout(self.gridLayout_5, 0, 0, 1, 1)
        self.TabWidget_MainGUI.addTab(self.HistoryData_page, "")

        self.gridLayout_4.addWidget(self.TabWidget_MainGUI, 0, 0, 1, 1)
        self.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(self)
        self.statusbar.setObjectName("statusbar")
        self.setStatusBar(self.statusbar)

        self.retranslateUi(self)
        self.TabWidget_MainGUI.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(self)

    def retranslateUi(self, MainWindow):

        _translate = QtCore.QCoreApplication.translate

        font = QFont()
        font.setFamily('Consolas')
        font.setPixelSize(20)

        MainWindow.setWindowTitle(_translate("MainWindow", "基于面部特征和头部姿态的卒中患者疲劳度检测"))
        self.groupBox_Player.setTitle(_translate("MainWindow", "实时检测画面"))
        self.groupBox_Player.setFont(font)
        self.groupBox_Qchart.setTitle(_translate("MainWindow", "疲劳特征曲线"))
        self.groupBox_Qchart.setFont(font)

        self.btn_save.setText(_translate("MainWindow", "保存数据"))
        self.btn_save.setFont(font)

        self.btn_cont.setText(_translate("MainWindow", "继  续"))
        self.btn_cont.setFont(font)

        # self.label_camera_show.setText(_translate("MainWindow", "实时检测画面展示"))
        self.btn_start.setText(_translate("MainWindow", "开  始"))
        self.btn_start.setFont(font)

        self.btn_stop.setText(_translate("MainWindow", "暂  停"))
        self.btn_stop.setFont(font)

        self.btn_camera.setText(_translate("MainWindow", "实时拍摄"))
        self.btn_camera.setFont(font)

        self.btn_select.setText(_translate("MainWindow", "本地文件"))
        self.btn_select.setFont(font)
        # self.label_fatigue_reminder.setText(_translate("MainWindow", "患者训练状态"))

        self.TabWidget_MainGUI.setTabText(self.TabWidget_MainGUI.indexOf(self.Camera_page),
                                          _translate("MainWindow", "实时疲劳检测页面"))
        self.TabWidget_MainGUI.setTabText(self.TabWidget_MainGUI.indexOf(self.HistoryData_page),
                                          _translate("MainWindow", "历史数据查询页面"))
        self.TabWidget_MainGUI.setFont(font)

        self.label_patientName.setText(_translate("MainWindow", "患者姓名"))
        self.label_patientName.setFont(font)
        self.label_trainNumber.setText(_translate("MainWindow", "训练次数"))
        self.label_trainNumber.setFont(font)
        self.btn_yes.setText(_translate("MainWindow", "确 定"))
        self.btn_yes.setFont(font)
        self.rbtn_roll.setText(_translate("MainWindow", "头部滚转角，阈值±30"))
        self.rbtn_roll.setFont(font)
        self.rbtn_pitch.setText(_translate("MainWindow", "头部俯仰角，阈值±12"))
        self.rbtn_pitch.setFont(font)
        self.rbtn_yaw.setText(_translate("MainWindow", "头部偏航角，阈值±18"))
        self.rbtn_yaw.setFont(font)
        self.rbtn_eye.setText(_translate("MainWindow", "眼部纵横比，阈值0.25"))
        self.rbtn_eye.setFont(font)
        self.rbtn_mouth.setText(_translate("MainWindow", "嘴部开合比，阈值0.75"))
        self.rbtn_mouth.setFont(font)

        self.label_patientName_show.setText(str(self.use_name))
        self.label_patientName_show.repaint()
        self.label_patientName_show.setFont(font)

    '''
    实时疲劳检测查询页面全部函数
    '''

    def clear(self):
        self.label_camera_show.clear()
        self.label_fatigue_reminder.clear()

        self.lineseries1.clear()
        self.lineseries2.clear()
        self.lineseries3.clear()
        self.lineseries4.clear()
        self.lineseries5.clear()

        self.count = 0

    def cameraVideo(self):
        self.clear()
        self.video_path = '0'
        self.getVideoSavePath()

    def selectVideo(self):
        self.clear()
        video_path_0 = QtWidgets.QFileDialog.getOpenFileName(None, "选择视频",
                                                             "D:/projects_of_python/StrokePatients_Fatigue_Detection/demo",
                                                             "Image files(*.mp4 *.avi)")  # 起始路径
        self.video_path = video_path_0[0]  # 获取视频名称
        if self.video_path == '':
            msy_box = QMessageBox(QMessageBox.Warning, "Warning", "请选择视频!")
            msy_box.exec_()
        else:
            self.selectVideoFirstFrame()

    def selectVideoFirstFrame(self):

        self.cap = cv2.VideoCapture(self.video_path)
        self.cap.set(1, 1)  # 取第一帧
        self.ret, self.frame = self.cap.read()

        if self.ret:
            self.frame = cv2.cvtColor(self.frame, cv2.COLOR_BGR2RGB)
            frameHeight, frameWidth, bytesPerComponent = self.frame.shape
            bytesPerLine = bytesPerComponent * frameWidth
            q_image = QImage(self.frame.data, frameWidth, frameHeight, bytesPerLine,
                             QImage.Format_RGB888).scaled(frameWidth * self.label_camera_show.height() / frameHeight,
                                                          self.label_camera_show.height())
            self.label_camera_show.setPixmap(QPixmap.fromImage(q_image))
            self.label_camera_show.setAlignment(Qt.AlignCenter)

    def start(self):
        self.clear()
        if self.video_path != '0':
            print("自选视频进行测试")
            self.cap = cv2.VideoCapture(self.video_path)
            self.fps = int(self.cap.get(cv2.CAP_PROP_FPS))

        else:
            print("选用摄像头进行测试")

            cameras = list_video_devices()
            print(f'======================  camera_list:{cameras}')
            idx = 0
            camera_id = cameras[idx][0]
            camera_name = cameras[idx][1]
            print(f'\n======================  preview camera:camera_id={camera_id} camera_name={camera_name}')
            self.cap = cv2.VideoCapture(camera_id, cv2.CAP_DSHOW)
            self.cap.set(cv2.CAP_PROP_FPS, 30)
            self.fps = int(self.cap.get(cv2.CAP_PROP_FPS))
            # self.timer_videoSave.start(self.fps)
            # self.timer_videoSave.timeout.connect(self.saveFrameImages)

            # self.cap = cv2.VideoCapture(0)
            # self.fps = int(self.cap.get(cv2.CAP_PROP_FPS))
            # self.timer_videoSave.start(self.fps)
            # self.timer_videoSave.timeout.connect(self.saveFrameImages)

            # if cv2.VideoCapture(1).grab():
            #     waitKey(100)
            #     print("外接相机可用")
            #     self.cap = cv2.VideoCapture(1, cv2.CAP_DSHOW)
            #     self.fps = int(self.cap.get(cv2.CAP_PROP_FPS))
            #     self.timer_videoSave.start(self.fps)
            #     self.timer_videoSave.timeout.connect(self.saveFrameImages)
            #
            # else:
            #     print("本机相机可用")
            #     self.cap = cv2.VideoCapture(0)
            #     self.fps = int(self.cap.get(cv2.CAP_PROP_FPS))
            #     self.timer_videoSave.start(self.fps)
            #     self.timer_videoSave.timeout.connect(self.saveFrameImages)

        self.Eye_det = Eye_Detect_Module.EyeDetector(show_processing=False)
        self.Head_pose = Pose_Estimation_Module.HeadPoseEstimator(show_axis=True)
        self.Scorer = Fatigue_Scorer_Module.AttentionScorer(self.fps, ear_threshold=0.25, ear_time_threshold=3,
                                                            perclos_threshold=0.1,
                                                            pitch_threshold=12, yaw_threshold=18, roll_threshold=30,
                                                            pose_time_threshold=4,
                                                            mouth_threshold=0.75, mouth_time_threshold=2, verbose=False)

        self.timer_camera.start(30)
        self.timer_camera.timeout.connect(self.openFrame)
        self.timer_lineChart.start(100)
        self.timer_lineChart.timeout.connect(self.change_linechart)

    def getVideoSavePath(self):

        video_path_all = os.path.abspath(os.path.join(os.getcwd(), "Fatigue_history_data"))
        self.video_path_use = os.path.join(video_path_all, self.use_name, "Video")

        if not os.path.exists(self.video_path_use):
            os.makedirs(self.video_path_use)
            self.Video_num = 0
            self.Video_name_list = None
            # print('当前文件夹不存在，已创建')
        else:
            self.Video_name_list = sorted(glob.glob(os.path.join(self.video_path_use, "*.mp4")))
            self.Video_num = len(self.Video_name_list)

        self.videoSavePath = self.video_path_use + '\\' + 'Video-' + str(self.Video_num + 1) + '.mp4'
        print(self.videoSavePath)

    def saveFrameImages(self):
        # print("保存视频帧",self.fps)
        self.frameCount += 1
        success, image = self.cap.read()
        cv2.imencode(".jpg", image)[1].tofile(os.path.join(self.video_path_use, f"frame-{self.frameCount:07}.jpg"))

    def openFrame(self):
        # print("处理视频帧")
        if self.video_path == "0":
            self.ret, self.frame = self.cap.read()
            if self.ret:
                self.frame = cv2.flip(self.frame, 2)
        else:
            self.ret, self.frame = self.cap.read()

        gray = cv2.cvtColor(self.frame, cv2.COLOR_BGR2GRAY)
        gray = cv2.bilateralFilter(gray, 5, 10, 10)

        faces = self.Detector(gray)
        if len(faces) > 0:
            self.tracker.start_track(self.frame, faces[0])
            self.tracker.update(self.frame)
            pos = self.tracker.get_position()
            cv2.rectangle(self.frame, (int(pos.left()), int(pos.top())),
                          (int(pos.right()), int(pos.bottom())), (0, 255, 0), 3)
            faces = sorted(faces, key=Utils.get_face_area, reverse=True)
            driver_face = faces[0]

            landmarks = self.Predictor(gray, driver_face)
            self.Eye_det.show_eye_keypoints(color_frame=self.frame, landmarks=landmarks)
            self.EAR = self.Eye_det.get_EAR(frame=gray, landmarks=landmarks)
            self.EAR = float('%.3f' % self.EAR)

            self.mouth_opening = self.Eye_det.get_mouth_Score(frame=gray, landmarks=landmarks)
            self.mouth_opening = float('%.3f' % self.mouth_opening)

            self.frame, self.roll, self.pitch, self.yaw = self.Head_pose.get_pose(frame=self.frame, landmarks=landmarks)
            self.roll = float('%.3f' % self.roll)
            self.pitch = float('%.3f' % self.pitch)
            self.yaw = float('%.3f' % self.yaw)

            self.asleep, self.tired, self.distracted = self.Scorer.eval_scores(self.EAR, self.mouth_opening, self.pitch,
                                                                               self.yaw, self.roll)

            self.fatigue_decide()

            self.DataListEye.append(self.EAR)
            self.DataListMouth.append(self.mouth_opening)
            self.DataListYaw.append(self.yaw)
            self.DataListRoll.append(self.roll)
            self.DataListPitch.append(self.pitch)

        self.frame = cv2.cvtColor(self.frame, cv2.COLOR_BGR2RGB)
        frameHeight, frameWidth, bytesPerComponent = self.frame.shape
        bytesPerLine = bytesPerComponent * frameWidth
        q_image = QImage(self.frame.data, frameWidth, frameHeight, bytesPerLine,
                         QImage.Format_RGB888).scaled(frameWidth * self.label_camera_show.height() / frameHeight,
                                                      self.label_camera_show.height())
        self.label_camera_show.setPixmap(QPixmap.fromImage(q_image))
        self.label_camera_show.setAlignment(Qt.AlignCenter)

    def fatigue_decide(self):
        label_fatigue = QFont()
        label_fatigue.setFamily("Times New Roman")
        label_fatigue.setBold(True)
        label_fatigue.setPixelSize(55)
        if self.tired == False and self.asleep == False and self.distracted == False:
            self.label_fatigue_reminder.setText("In Normal Training!")
            self.label_fatigue_reminder.setStyleSheet("color:#06d506")
        if self.distracted:
            self.label_fatigue_reminder.setText("The patient is distracted!")
            self.label_fatigue_reminder.setStyleSheet("color:#f7b776")
        if self.tired:
            self.label_fatigue_reminder.setText("The patient is tired!")
            self.label_fatigue_reminder.setStyleSheet("color:#ff9900")
        if self.asleep:
            self.label_fatigue_reminder.setText("The patient is asleep!")
            self.label_fatigue_reminder.setStyleSheet("color:#df3939")

        self.label_fatigue_reminder.setAlignment(Qt.AlignCenter)
        self.label_fatigue_reminder.setFont(label_fatigue)

    def stop(self):
        self.timer_camera.stop()
        self.timer_lineChart.stop()
        if self.video_path == '0':
            self.timer_videoSave.stop()

    def contiPlay(self):
        self.timer_camera.start()
        self.timer_lineChart.start()
        if self.video_path == '0':
            self.timer_videoSave.start()

    def draw_linechart_eye(self, data):
        self.count += 1
        self.lineseries1.append(self.count, data * 10)
        self.lineseries1.setName("眼部纵横比")
        self.lineseries1.setVisible(True)
        self.linechart1.addSeries(self.lineseries1)

        if self.count < 150:
            self.lineaxisX1.setRange(0, 200)
        else:
            self.lineaxisX1.setRange(self.count - 150, self.count + 50)
        self.lineaxisX1.setTickCount(10)
        self.lineaxisX1.setTitleText("时间")

        self.lineaxisY1.setRange(0, 10)
        self.lineaxisY1.setTickCount(5)
        self.lineaxisY1.setTitleText("EAR")

        self.lineaxisY1.setGridLineVisible(True)
        self.lineaxisY1.setGridLineColor(Qt.gray)
        self.lineaxisX1.setGridLineVisible(True)
        self.lineaxisX1.setGridLineColor(Qt.gray)

        self.linechart1.addAxis(self.lineaxisX1, Qt.AlignBottom)
        self.linechart1.addAxis(self.lineaxisY1, Qt.AlignLeft)

        self.lineseries1.attachAxis(self.lineaxisX1)
        self.lineseries1.attachAxis(self.lineaxisY1)

        self.linechart1.setVisible(True)
        self.graphicsView_eye.setChart(self.linechart1)

    def draw_linechart_mouth(self, data):
        # self.count += 1
        self.lineseries2.append(self.count, data * 10)
        self.lineseries2.setName("嘴部开合比")
        self.lineseries2.setVisible(True)
        self.linechart2.addSeries(self.lineseries2)

        if self.count < 150:
            self.lineaxisX2.setRange(0, 200)
        else:
            self.lineaxisX2.setRange(self.count - 150, self.count + 50)
        self.lineaxisX2.setTickCount(10)
        self.lineaxisX2.setTitleText("时间")

        self.lineaxisY2.setRange(0, 15)
        self.lineaxisY2.setTickCount(5)
        self.lineaxisY2.setTitleText("MAR")
        self.lineaxisY2.setGridLineVisible(True)
        self.lineaxisY2.setGridLineColor(Qt.gray)
        self.lineaxisX2.setGridLineVisible(True)
        self.lineaxisX2.setGridLineColor(Qt.gray)

        self.linechart2.addAxis(self.lineaxisX2, Qt.AlignBottom)
        self.linechart2.addAxis(self.lineaxisY2, Qt.AlignLeft)

        self.lineseries2.attachAxis(self.lineaxisX2)
        self.lineseries2.attachAxis(self.lineaxisY2)

        self.linechart2.setVisible(True)
        self.graphicsView_mouth.setChart(self.linechart2)

    def draw_linechart_yaw(self, data):
        # self.count += 1
        self.lineseries3.append(self.count, data)
        self.lineseries3.setName("头部偏航角")
        self.lineseries3.setVisible(True)
        self.linechart3.addSeries(self.lineseries3)

        if self.count < 150:
            self.lineaxisX3.setRange(0, 200)
        else:
            self.lineaxisX3.setRange(self.count - 150, self.count + 50)
        self.lineaxisX3.setTickCount(10)
        self.lineaxisX3.setTitleText("时间")

        self.lineaxisY3.setRange(0, 50)
        self.lineaxisY3.setTickCount(5)
        self.lineaxisY3.setTitleText("Yaw")
        self.lineaxisY3.setGridLineVisible(True)
        self.lineaxisY3.setGridLineColor(Qt.gray)
        self.lineaxisX3.setGridLineVisible(True)
        self.lineaxisX3.setGridLineColor(Qt.gray)

        self.linechart3.addAxis(self.lineaxisX3, Qt.AlignBottom)
        self.linechart3.addAxis(self.lineaxisY3, Qt.AlignLeft)

        self.lineseries3.attachAxis(self.lineaxisX3)
        self.lineseries3.attachAxis(self.lineaxisY3)

        self.linechart3.setVisible(True)
        self.graphicsView_yaw.setChart(self.linechart3)

    def draw_linechart_roll(self, data):
        # self.count += 1
        self.lineseries4.append(self.count, data)
        self.lineseries4.setName("头部滚转角")
        self.lineseries4.setVisible(True)
        self.linechart4.addSeries(self.lineseries4)

        if self.count < 150:
            self.lineaxisX4.setRange(0, 200)
        else:
            self.lineaxisX4.setRange(self.count - 150, self.count + 50)
        self.lineaxisX4.setTickCount(10)
        self.lineaxisX4.setTitleText("时间")

        self.lineaxisY4.setRange(0, 50)
        self.lineaxisY4.setTickCount(5)
        self.lineaxisY4.setTitleText("Roll")
        self.lineaxisY4.setGridLineVisible(True)
        self.lineaxisY4.setGridLineColor(Qt.gray)
        self.lineaxisX4.setGridLineVisible(True)
        self.lineaxisX4.setGridLineColor(Qt.gray)

        self.linechart4.addAxis(self.lineaxisX4, Qt.AlignBottom)
        self.linechart4.addAxis(self.lineaxisY4, Qt.AlignLeft)

        self.lineseries4.attachAxis(self.lineaxisX4)
        self.lineseries4.attachAxis(self.lineaxisY4)

        self.linechart4.setVisible(True)
        self.graphicsView_roll.setChart(self.linechart4)

    def draw_linechart_pitch(self, data):
        # self.count += 1
        self.lineseries5.append(self.count, data)
        self.lineseries5.setName("头部俯仰角")
        self.lineseries5.setVisible(True)
        self.linechart5.addSeries(self.lineseries5)

        if self.count < 150:
            self.lineaxisX5.setRange(0, 200)
        else:
            self.lineaxisX5.setRange(self.count - 150, self.count + 50)
        self.lineaxisX5.setTickCount(10)
        self.lineaxisX5.setTitleText("时间")

        self.lineaxisY5.setRange(0, 30)
        self.lineaxisY5.setTickCount(5)
        self.lineaxisY5.setTitleText("Pitch")
        self.lineaxisY5.setGridLineVisible(True)
        self.lineaxisY5.setGridLineColor(Qt.gray)
        self.lineaxisX5.setGridLineVisible(True)
        self.lineaxisX5.setGridLineColor(Qt.gray)

        self.linechart5.addAxis(self.lineaxisX5, Qt.AlignBottom)
        self.linechart5.addAxis(self.lineaxisY5, Qt.AlignLeft)

        self.lineseries5.attachAxis(self.lineaxisX5)
        self.lineseries5.attachAxis(self.lineaxisY5)

        self.linechart5.setVisible(True)
        self.graphicsView_pitch.setChart(self.linechart5)

    def change_linechart(self):
        Data_eye = self.EAR
        self.draw_linechart_eye(Data_eye)

        Data_mouth = self.mouth_opening
        self.draw_linechart_mouth(Data_mouth)

        Data_yaw = self.yaw
        Data_yaw = abs(Data_yaw)
        self.draw_linechart_yaw(Data_yaw)

        Data_roll = self.roll
        Data_roll = abs(Data_roll)
        self.draw_linechart_roll(Data_roll)

        Data_pitch = self.pitch
        Data_pitch = abs(Data_pitch)
        self.draw_linechart_pitch(Data_pitch)

    def get_patient_name(self):
        file_path_all = os.path.abspath(os.path.join(os.getcwd(), "Fatigue_history_data"))
        self.file_path_use = os.path.join(file_path_all, self.use_name, "Excel")

        if not os.path.exists(self.file_path_use):
            os.makedirs(self.file_path_use)
            self.Excel_num = 0
            self.Excel_name_list = None
            # print('当前文件夹不存在，已创建')
        else:
            self.Excel_name_list = os.listdir(self.file_path_use)
            self.Excel_num = len(self.Excel_name_list)
            # print('当前文件夹存在，且有',self.Excel_num,'个文件')

    def generateVideo(self, inp: str, out: str):

        img_array = []
        images_list = sorted(glob.glob(os.path.join(inp, "frame-*.jpg")))

        with open(images_list[0], "rb") as f:
            img = Image.open(f)
            img = img.convert("RGB")
            size = (img.width, img.height)
            img_array.append(cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR))

        for filename in tqdm(images_list[1:]):
            with open(filename, "rb") as f:
                img = Image.open(f)
                img = img.convert("RGB")
                img_array.append(cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR))

        outWriter = cv2.VideoWriter(out, self.FOURCC, self.fps, size, )

        for i in range(len(img_array)):
            outWriter.write(img_array[i])
        outWriter.release()
        cv2.destroyAllWindows()

        for img in images_list:
            os.remove(img)
        print("图像删除完毕")

    def save(self):
        self.get_patient_name()

        file_path_data = self.file_path_use + '\\' + '第' + str(self.Excel_num + 1) + '次训练数据' + '.xlsx'
        dataExcel = xlsxwriter.Workbook(file_path_data)
        dataSheet = dataExcel.add_worksheet('五类疲劳数据')
        print(file_path_data)
        listLen = len(self.DataListEye)
        print(listLen)
        for i in range(0, listLen):
            dataSheet.write(i, 0, self.DataListEye[i])
            dataSheet.write(i, 1, self.DataListMouth[i])
            dataSheet.write(i, 2, self.DataListYaw[i])
            dataSheet.write(i, 3, self.DataListRoll[i])
            dataSheet.write(i, 4, self.DataListPitch[i])
        dataExcel.close()

        if self.video_path == "0":
            self.generateVideo(self.video_path_use, self.videoSavePath)
            self.frameCount = 0

        QMessageBox.information(self, 'Successfully', '数据已保存成功！')

    '''
    历史数据查询页面全部函数
    '''

    def renewComboBox(self):
        # print("点击下拉框")
        self.get_patient_name()
        if self.Excel_name_list is None:
            QMessageBox.information(self, 'Sorry', '暂无历史训练数据！', QMessageBox.Ok)
        else:
            self.comboBox.clear()
            self.comboBox.addItems(self.Excel_name_list)
        self.comboBox.currentTextChanged.connect(self.getComboBoxCurrentText)

    def getComboBoxCurrentText(self):
        # print("获取选择的内容")
        selectExcel = self.comboBox.currentText()
        self.selectExcelPath = os.path.join(self.file_path_use, selectExcel)
        # print(self.selectExcelPath)

    def showHisFatigueChart(self):
        print("点击确定")
        print(self.selectExcelPath)
        path = self.selectExcelPath
        # xlsx = pandas.read_excel(path,engine='openpyxl')
        xlsx = openpyxl.load_workbook(path)
        sheet = xlsx.worksheets[0]
        rows = sheet.max_row

        self.ExcelEye = []
        for i in range(1, rows + 1):
            cell_value = sheet.cell(row=i, column=1).value
            self.ExcelEye.append(cell_value)
        self.ExcelMouth = []
        for i in range(1, rows + 1):
            cell_value = sheet.cell(row=i, column=2).value
            self.ExcelMouth.append(cell_value)
        self.ExcelYaw = []
        for i in range(1, rows + 1):
            cell_value = sheet.cell(row=i, column=3).value
            self.ExcelYaw.append(cell_value)
        self.ExcelRoll = []
        for i in range(1, rows + 1):
            cell_value = sheet.cell(row=i, column=4).value
            self.ExcelRoll.append(cell_value)
        self.ExcelPitch = []
        for i in range(1, rows + 1):
            cell_value = sheet.cell(row=i, column=5).value
            self.ExcelPitch.append(cell_value)

        # self.ExcelEye = xlsx.values[:, 0]
        # self.ExcelMouth = xlsx.values[:, 1]
        # self.ExcelYaw = xlsx.values[:, 2]
        # self.ExcelRoll = xlsx.values[:, 3]
        # self.ExcelPitch = xlsx.values[:, 4]

    def draw(self, data, cls):
        self.line.axes.clear()
        self.line.fig.canvas.draw_idle()

        num = len(data)
        xdata = list(range(num))

        if cls == 1:
            self.threshold = 0.25
            ydata_1 = [self.threshold] * num
            ydata_2 = ydata_1

        elif cls == 2:
            self.threshold = 0.75
            ydata_1 = [self.threshold] * num
            ydata_2 = ydata_1
        elif cls == 3:
            self.threshold = 18
            ydata_1 = [self.threshold] * num
            ydata_2 = [-self.threshold] * num
        elif cls == 4:
            self.threshold = 30
            ydata_1 = [self.threshold] * num
            ydata_2 = [-self.threshold] * num
        else:
            self.threshold = 12
            ydata_1 = [self.threshold] * num
            ydata_2 = [-self.threshold] * num

        self.line.axes.plot(xdata, data)
        self.line.axes.plot(xdata, ydata_1)
        self.line.axes.plot(xdata, ydata_2)

        self.layout.addWidget(self.toolbar, Qt.AlignTop)
        self.layout.addWidget(self.line)
        self.widget.setLayout(self.layout)
        self.widget.showFullScreen()

    def widget_show_eye(self):
        print('eye')
        cls = 1
        self.draw(self.ExcelEye, cls)

    def widget_show_mouth(self):
        print('mouth')
        cls = 2
        self.draw(self.ExcelMouth, cls)

    def widget_show_yaw(self):
        print('yaw')
        cls = 3
        self.draw(self.ExcelYaw, cls)

    def widget_show_roll(self):
        print('roll')
        cls = 4
        self.draw(self.ExcelRoll, cls)

    def widget_show_pitch(self):
        print('pitch')
        cls = 5
        self.draw(self.ExcelPitch, cls)


# 自定义ComboBox
class MyComboBox(QComboBox):
    clicked = pyqtSignal()

    def showPopup(self):
        self.clicked.emit()
        super(MyComboBox, self).showPopup()  # 调用父类的showPopup()


# 自定义Canvas
class MyCanvas(FigureCanvasQTAgg):
    def __init__(self, parent=None, width=8, height=4, dpi=100):
        self.fig = Figure(figsize=(width, height))
        self.axes = self.fig.add_subplot(111)
        self.fig.supxlabel('Time')
        self.fig.supylabel('Value')

        FigureCanvasQTAgg.__init__(self, self.fig)
        FigureCanvasQTAgg.setSizePolicy(self,
                                        QtWidgets.QSizePolicy.Expanding,
                                        QtWidgets.QSizePolicy.Expanding)
        FigureCanvasQTAgg.updateGeometry(self)

#
#
# if __name__ == '__main__':
#     app = QApplication(sys.argv)
#     window = Main()
#     window.show()
#     sys.exit(app.exec_())
